# Fantasy premier league task
# Imports
import requests
import json
import openpyxl
import os
from datetime import datetime

# Variables
users = [86218]
startingColor = openpyxl.styles.PatternFill(start_color='4DED30', end_color='4DED30', fill_type='solid');
benchColor = openpyxl.styles.PatternFill(start_color='B5651D', end_color='B5651D', fill_type='solid');
inElementColor = openpyxl.styles.PatternFill(start_color='26D701', end_color='26D701', fill_type='solid');
outElementColor = openpyxl.styles.PatternFill(start_color='FF2400', end_color='FF2400', fill_type='solid');

# Create new directory for runtime
newpath = r'./' + str(datetime.now().date()) 
if not os.path.exists(newpath):
    os.makedirs(newpath)


# Create all players dictionary
ALL_PLAYERS = 'https://fantasy.premierleague.com/api/bootstrap-static/'
players = {}
r = requests.get(ALL_PLAYERS)
json_resp = json.loads(r.text)
for elem in json_resp['elements']:
	players[elem['id']] = str(elem['first_name']) + ' ' + str(elem['second_name'])


# For each user defined
for user in users:
	####################################
	#         Workbook setup           #
	####################################
	# Make a new workboob
	wb = openpyxl.Workbook();

	# Create new sheet for the user
	ws = wb.active
	ws.title = 'Picks'
	
	# Set cell dimensions
	ws.column_dimensions['A'].width = 30;
	ws.column_dimensions['B'].width = 50;
	ws.column_dimensions['C'].width = 30;
	sheet_row = 1

	###################################
	#   Get picks for each game week  #
	###################################
	event = 1
	s = 0
	r = requests.get('https://fantasy.premierleague.com/api/entry/' + str(user) + '/event/' + str(event) +'/picks/')
	json_resp = json.loads(r.text)
	while ('detail' not in json_resp):
		# Write the first entry
		ws['A' + str(sheet_row)] = 'Gameweek ' + str(event)
		ws['B' + str(sheet_row)] = 'Player IDs(Green starting, brown bench)'
		ws['C' + str(sheet_row)] = 'Total points captured'
		sheet_row += 1
		ws['C' + str(sheet_row)] = json_resp['entry_history']['points']
		for elem in json_resp['picks']:
			ws['B' + str(sheet_row)] = players[elem['element']]
			if (elem['multiplier'] >= 1):
				ws['B' + str(sheet_row)].fill = startingColor
			else:
				ws['B' + str(sheet_row)].fill = benchColor
			sheet_row += 1
		
		event += 1
		s  = 0
		r = requests.get('https://fantasy.premierleague.com/api/entry/' + str(user) + '/event/' + str(event) +'/picks/')
		json_resp = json.loads(r.text)


	################################
	#    Get last week transfers   #
	################################
	events_so_far = event - 1
	ws = wb.create_sheet('Transfers')
	sheet_row = 1
	ws.column_dimensions['A'].width = 50;
	ws.column_dimensions['B'].width = 50;
	ws['A' + str(sheet_row)] = 'Transfers during gameweek ' + str(events_so_far)
	sheet_row += 1
	ws['A' + str(sheet_row)] = 'Element in'
	ws['B' + str(sheet_row)] = 'Element out'
	sheet_row += 1

	r = requests.get('https://fantasy.premierleague.com/api/entry/' + str(user) + '/transfers/')
	json_resp = json.loads(r.text)
	for elem in json_resp:
		if (elem['event'] == events_so_far):
			ws['A' + str(sheet_row)] = elem['element_in']
			ws['A' + str(sheet_row)].fill = inElementColor
			ws['B' + str(sheet_row)] = elem['element_out']
			ws['B' + str(sheet_row)].fill = outElementColor


	wb.save(newpath + '/' + 'User_' + str(user) + '_requested_at_' + str(datetime.now().date())  + '.xlsx');
